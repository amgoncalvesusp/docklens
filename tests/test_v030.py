"""Contract tests for the DockLens 0.3 export and result model."""

from __future__ import annotations

from dataclasses import FrozenInstanceError, replace
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from docklens import batch_runner as br
from docklens import export
from docklens.interaction_core import Atom, compute_interactions
from docklens.interaction_core import CUTOFFS
from docklens.results import Endpoint


def test_pose_identity_and_immutable_key_recompute(fixture_path):
    result = br.run([fixture_path("two_pose_complex_sol7.pdb")])

    assert len(result.summaries) == 2
    assert len({summary.pose_id for summary in result.summaries}) == 2
    assert all(detail.pose_id for detail in result.details)
    assert {detail.pose_id for detail in result.details} <= {
        summary.pose_id for summary in result.summaries
    }

    updated = br.recompute_key(result, ["SER70A"])
    assert updated is not result
    assert result.key_residues == frozenset()
    assert updated.key_residues == frozenset({"SER70A"})
    assert updated.parameters.key_residues == ("SER70A",)
    with pytest.raises(FrozenInstanceError):
        updated.summaries[0].pose = 99


def test_parse_failure_is_auditable_and_does_not_drop_valid_input(fixture_path):
    result = br.run(
        [fixture_path("invalid.pdb"), fixture_path("minimal_complex.pdb")]
    )

    assert result.summaries
    statuses = {record.status for record in result.input_qc}
    assert "error" in statuses
    assert "success" in statuses
    assert any(record.message for record in result.input_qc if record.status == "error")


def test_residue_matrix_uses_pose_residue_and_type(fixture_path):
    result = br.run([fixture_path("two_pose_complex_sol7.pdb")])

    matrix = export.residue_matrix_dataframe(result, mode="count")
    assert len(matrix) == 2
    assert matrix.columns.nlevels == 2
    assert ("SER70A", "hbond") in matrix.columns
    numeric = matrix.drop(columns="Identity", level=0)
    assert int(numeric.to_numpy().sum()) == len(result.details)

    presence = export.residue_matrix_dataframe(result, mode="presence")
    values = presence.drop(columns="Identity", level=0).to_numpy().ravel()
    assert set(values) <= {0, 1}


def test_water_bridge_is_one_semantic_record():
    receptor = Atom(0, "O", "OG", "SER", "70", "A", coord=(3.0, 0.0, 0.0), serial=1)
    ligand = Atom(1, "N", "N1", "LIG", "1", "B", coord=(-0.52, 2.95, 0.0), serial=2)
    water = Atom(2, "O", "O", "HOH", "8", "W", coord=(0.0, 0.0, 0.0), serial=3)
    receptor.side, ligand.side, water.side = "receptor", "ligand", "water"

    records = compute_interactions([receptor], [ligand], [water], ["water_bridge"])

    assert len(records) == 1
    assert records[0]["a_obj"] is receptor
    assert records[0]["b_obj"] is ligand
    assert records[0]["water_obj"] is water
    assert records[0]["a_role"] == "donor_acceptor"
    assert records[0]["b_role"] == "donor_acceptor"
    assert records[0]["receptor_water_distance"] == pytest.approx(3.0)
    assert records[0]["ligand_water_distance"] == pytest.approx(3.0, abs=0.01)


def test_xlsx_v2_has_five_sheets_styles_qc_and_safe_text(fixture_path, tmp_path):
    result = br.run(
        [fixture_path("minimal_complex.pdb"), fixture_path("invalid.pdb")]
    )
    label = "=HYPERLINK(\"https://invalid\")"
    dangerous = replace(
        result,
        details=tuple(replace(item, source_file=label) for item in result.details),
        summaries=tuple(replace(item, source_file=label) for item in result.summaries),
        input_qc=tuple(replace(item, source_file=label) for item in result.input_qc),
    )
    output = Path(export.export_xlsx(dangerous, tmp_path / "docklens.xlsx"))

    workbook = openpyxl.load_workbook(output, data_only=False)
    assert workbook.sheetnames == [
        "Summary",
        "Residue Matrix",
        "Detail",
        "Parameters",
        "Input QC",
    ]
    assert workbook["Summary"].freeze_panes
    assert workbook["Detail"].auto_filter.ref
    assert workbook["Residue Matrix"].merged_cells.ranges
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )


def test_filtered_export_recomputes_all_derived_views(fixture_path):
    result = br.run([fixture_path("minimal_complex.pdb")])
    selected_type = result.details[0].interaction_type
    view = export.build_export_view(
        result,
        br.ExportFilter(scope="filtered", interaction_types=frozenset({selected_type})),
    )

    assert view.details
    assert {detail.interaction_type for detail in view.details} == {selected_type}
    assert sum(summary.n_total_interactions for summary in view.summaries) == len(
        view.details
    )


def test_filtered_export_preserves_poses_with_zero_counts(fixture_path):
    result = br.run([fixture_path("two_pose_complex_sol7.pdb")])

    view = export.build_export_view(
        result,
        br.ExportFilter(scope="filtered", interaction_types=frozenset()),
    )

    assert not view.details
    assert len(view.summaries) == len(result.summaries)
    assert all(summary.n_total_interactions == 0 for summary in view.summaries)
    assert all(not any(summary.counts.values()) for summary in view.summaries)
    assert len(export.residue_matrix_dataframe(view)) == len(result.summaries)


def test_export_views_include_source_paths_and_endpoint_elements(fixture_path):
    result = br.run([fixture_path("minimal_complex.pdb")])

    summary = export.summary_dataframe(result)
    detail = export.detail_dataframe(result)
    qc = export.input_qc_dataframe(result)

    assert summary.loc[0, "source_path"] == result.summaries[0].source_path
    assert detail.loc[0, "source_path"] == result.details[0].source_path
    assert detail.loc[0, "ligand_element"] == result.details[0].ligand.element
    assert detail.loc[0, "receptor_element"] == result.details[0].receptor.element
    assert "water_element" in detail.columns
    assert qc.loc[0, "source_path"] == result.input_qc[0].source_path


def test_export_text_filter_matches_visible_ui_fields_only(fixture_path):
    result = br.run([fixture_path("minimal_complex.pdb")])
    pose_id = result.details[0].pose_id

    view = export.build_export_view(
        result, br.ExportFilter(scope="filtered", text=pose_id)
    )

    assert not view.details
    assert len(view.summaries) == len(result.summaries)


def test_xlsx_splits_wide_matrix_and_uses_filterable_second_header(
    fixture_path, tmp_path, monkeypatch
):
    result = br.run([fixture_path("minimal_complex.pdb")])
    original = result.details[0]
    details = tuple(
        replace(
            original,
            interaction_id="%s:W%03d" % (original.pose_id, index),
            receptor=replace(original.receptor, resseq=str(100 + index)),
        )
        for index in range(6)
    )
    result = replace(result, details=details)
    monkeypatch.setattr(export, "_MAX_XLSX_COLUMNS", 10)

    output = export.export_xlsx(result, tmp_path / "wide.xlsx")
    workbook = openpyxl.load_workbook(output)
    matrix_names = [name for name in workbook.sheetnames if name.startswith("Residue Matrix")]

    assert matrix_names == ["Residue Matrix 1", "Residue Matrix 2", "Residue Matrix 3"]
    for name in matrix_names:
        sheet = workbook[name]
        assert sheet.max_column <= 10
        assert sheet.auto_filter.ref.startswith("A2:")
        assert [sheet.cell(2, column).value for column in range(1, 9)] == [
            "ligand_id",
            "source_file",
            "source_path",
            "source_id",
            "pose_id",
            "sol",
            "pose",
            "docking_score",
        ]


def test_xlsx_rejects_views_beyond_excel_row_limit(fixture_path, tmp_path, monkeypatch):
    result = br.run([fixture_path("minimal_complex.pdb")])
    monkeypatch.setattr(export, "_MAX_XLSX_ROWS", 1)

    with pytest.raises(ValueError, match="row limit"):
        export.export_xlsx(result, tmp_path / "too-many-rows.xlsx")


def test_matrix_row_limit_includes_legend_spacing(monkeypatch):
    matrix = pd.DataFrame([[1]])
    highest_written_row = len(matrix) + 4 + len(export.VALID_TYPES)
    monkeypatch.setattr(export, "_MAX_XLSX_ROWS", highest_written_row - 1)

    with pytest.raises(ValueError, match="Residue Matrix.*row limit"):
        export._validate_xlsx_rows({}, matrix)


def test_sanitize_dataframe_handles_pandas_string_dtype():
    frame = pd.DataFrame(
        {"text": pd.Series(["=1+1", "safe"], dtype="string"), "number": [1, 2]}
    )

    sanitized = export._sanitize_dataframe(frame)

    assert sanitized.loc[0, "text"] == "'=1+1"
    assert sanitized.loc[1, "text"] == "safe"
    assert sanitized["number"].tolist() == [1, 2]


@pytest.mark.parametrize("prefix", ["=1+1", "+1+1", "-1+1", "@SUM(A1)", "  =1+1"])
def test_csv_neutralizes_all_formula_prefixes(fixture_path, tmp_path, prefix):
    result = br.run([fixture_path("minimal_complex.pdb")])
    result = replace(
        result,
        summaries=tuple(replace(item, source_file=prefix) for item in result.summaries),
        details=tuple(replace(item, source_file=prefix) for item in result.details),
    )

    summary_path, detail_path = export.export_csv(result, tmp_path / "safe")

    assert pd.read_csv(summary_path).loc[0, "source_file"].startswith("'")
    assert pd.read_csv(detail_path).loc[0, "source_file"].startswith("'")


def test_water_bridge_fields_are_serialized_into_detail_and_matrix(
    fixture_path, tmp_path
):
    result = br.run([fixture_path("minimal_complex.pdb")])
    original = result.details[0]
    bridge = replace(
        original,
        interaction_type="water_bridge",
        water=Endpoint("water", "atom", "O", (99,), "HOH", "8", "W", "O", "bridge"),
        receptor_water_distance_A=2.8,
        ligand_water_distance_A=3.0,
        water_angle_deg=110.0,
    )
    result = replace(result, details=(bridge,))

    detail = export.detail_dataframe(result)
    matrix = export.residue_matrix_dataframe(result)

    assert detail.loc[0, "water_residue"] == "HOH8W"
    assert detail.loc[0, "water_element"] == "O"
    assert detail.loc[0, "water_angle_deg"] == 110.0
    assert int(matrix[(bridge.receptor_residue, "water_bridge")].iloc[0]) == 1


def test_qc_only_workbook_is_exportable(fixture_path, tmp_path):
    result = br.run([fixture_path("invalid.pdb")])

    output = export.export_xlsx(result, tmp_path / "qc-only.xlsx")
    workbook = openpyxl.load_workbook(output, read_only=True)

    assert not result.summaries
    assert workbook.sheetnames == [
        "Summary",
        "Residue Matrix",
        "Detail",
        "Parameters",
        "Input QC",
    ]
    assert workbook["Input QC"].max_row == 2
    workbook.close()


def test_csv_and_xlsx_accept_pathlike(fixture_path, tmp_path):
    result = br.run([fixture_path("minimal_complex.pdb")])

    paths = export.export_csv(result, tmp_path / "out")
    assert all(Path(path).is_file() for path in paths)
    assert Path(export.export_xlsx(result, tmp_path / "out.xlsx")).is_file()


def test_run_captures_cutoffs_without_mutating_global_state(fixture_path):
    original = dict(CUTOFFS)

    result = br.run([fixture_path("minimal_complex.pdb")], hbond_preset="dsv")

    assert dict(result.parameters.cutoffs)["hbond_dist"] == 3.5
    assert dict(CUTOFFS) == original
