"""Non-visual packaged-application smoke check used by release builds."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from . import __version__
from .export import export_xlsx
from .results import AnalysisParameters, make_result


def run_self_check():
    parameters = AnalysisParameters(app_version=__version__)
    result = make_result(parameters=parameters)
    with tempfile.TemporaryDirectory(prefix="docklens-self-check-") as directory:
        output = Path(export_xlsx(result, Path(directory) / "self-check.xlsx"))
        workbook = load_workbook(output, read_only=True)
        expected = [
            "Summary",
            "Residue Matrix",
            "Key Residue Coverage",
            "Detail",
            "Parameters",
            "Input QC",
        ]
        actual = workbook.sheetnames
        workbook.close()
        if actual != expected:
            raise RuntimeError("Unexpected workbook schema: %r" % actual)

    # Exercise the packaged Qt runtime and real main window last, then return
    # normally so the one-file bootloader can reap its child and clean up.
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PyQt5 import QtCore, QtWidgets

    from .main_window import MainWindow

    existing_app = QtWidgets.QApplication.instance()
    application = existing_app or QtWidgets.QApplication(["DockLens", "--self-check"])
    window = MainWindow()
    window.show()
    QtCore.QTimer.singleShot(0, window.close)
    QtCore.QTimer.singleShot(0, application.quit)
    application.exec_()
    return 0
