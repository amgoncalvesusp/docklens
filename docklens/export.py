"""CSV/XLSX export with residue matrices, provenance and input QC."""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .export_views import (
    DETAIL_BASE_COLS,
    SUMMARY_BASE_COLS,
    build_export_view,
    detail_dataframe,
    input_qc_dataframe,
    parameters_dataframe,
    residue_matrix_dataframe,
    summary_dataframe,
)
from .interaction_core import INTERACTION_COLORS, VALID_TYPES, color_hex
from .results import ExportFilter

DETAIL_COLS = DETAIL_BASE_COLS
SUMMARY_BASE_COLS = SUMMARY_BASE_COLS
_INVALID_XML = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MAX_CELL_TEXT = 32767
_MAX_XLSX_COLUMNS = 16384
_MAX_XLSX_ROWS = 1048576


def _sanitize_cell(value):
    if not isinstance(value, str):
        return value
    value = _INVALID_XML.sub("", value)
    if len(value) > _MAX_CELL_TEXT:
        value = value[: _MAX_CELL_TEXT - 1] + "…"
    significant = value.lstrip(" \t\r\n")
    if significant.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _sanitize_dataframe(frame):
    result = frame.copy()
    for column in result.columns:
        result[column] = result[column].map(_sanitize_cell)
    return result


def _atomic_csv(frame, path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", prefix=".docklens-", dir=path.parent,
        delete=False, encoding="utf-8-sig", newline=""
    )
    tmp_path = Path(handle.name)
    handle.close()
    try:
        _sanitize_dataframe(frame).to_csv(tmp_path, index=False, encoding="utf-8-sig")
        os.replace(tmp_path, path)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def export_csv(result, path_prefix, export_filter=None) -> list:
    """Write compatible Summary/Detail CSV files atomically."""
    export_filter = export_filter or ExportFilter()
    view = build_export_view(result, export_filter)
    prefix = os.fspath(path_prefix)
    if prefix.lower().endswith(".csv"):
        prefix = prefix[:-4]
    summary_path = Path(prefix + "_summary.csv")
    detail_path = Path(prefix + "_detail.csv")
    _atomic_csv(summary_dataframe(view), summary_path)
    _atomic_csv(detail_dataframe(view), detail_path)
    return [str(summary_path), str(detail_path)]


def _fill(hex_color):
    argb = "FF" + hex_color.lstrip("#").upper()
    return PatternFill(start_color=argb, end_color=argb, fill_type="solid")


def _write_matrix_sheet(workbook, matrix, title="Residue Matrix", position=1):
    worksheet = workbook.create_sheet(title, position)
    for column_index, (group, label) in enumerate(matrix.columns, 1):
        worksheet.cell(1, column_index, _sanitize_cell(group))
        worksheet.cell(2, column_index, _sanitize_cell(label))
        if label in INTERACTION_COLORS:
            worksheet.cell(2, column_index).fill = _fill(color_hex(label))
    start = 1
    while start <= matrix.shape[1]:
        group = matrix.columns[start - 1][0]
        end = start
        while end < matrix.shape[1] and matrix.columns[end][0] == group:
            end += 1
        if end > start:
            worksheet.merge_cells(
                start_row=1, start_column=start, end_row=1, end_column=end
            )
        worksheet.cell(1, start).alignment = Alignment(horizontal="center")
        start = end + 1
    for row_index, row in enumerate(matrix.itertuples(index=False, name=None), 3):
        for column_index, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            worksheet.cell(row_index, column_index, _sanitize_cell(value))
    identity_columns = sum(group == "Identity" for group, _label in matrix.columns)
    worksheet.freeze_panes = "%s3" % get_column_letter(identity_columns + 1)
    worksheet.auto_filter.ref = "A2:%s%d" % (
        get_column_letter(worksheet.max_column),
        worksheet.max_row,
    )
    for column in range(1, worksheet.max_column + 1):
        values = [worksheet.cell(row, column).value for row in range(2, min(worksheet.max_row, 200) + 1)]
        width = max((len(str(value)) for value in values if value is not None), default=8)
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 42)
    legend_row = max(worksheet.max_row + 2, 5)
    worksheet.cell(legend_row, 1, "Legend")
    for offset, kind in enumerate(VALID_TYPES, 1):
        cell = worksheet.cell(legend_row + offset, 1, kind)
        cell.fill = _fill(color_hex(kind))
    return worksheet


def _matrix_chunks(matrix):
    """Split only residue columns while repeating pose identity in every sheet."""
    identity = [
        index for index, (group, _label) in enumerate(matrix.columns)
        if group == "Identity"
    ]
    values = [index for index in range(matrix.shape[1]) if index not in identity]
    capacity = _MAX_XLSX_COLUMNS - len(identity)
    if capacity < 1 and values:
        raise ValueError("Residue Matrix identity columns exceed Excel's column limit.")
    if matrix.shape[1] <= _MAX_XLSX_COLUMNS:
        return [matrix]
    return [
        matrix.iloc[:, identity + values[start : start + capacity]]
        for start in range(0, len(values), capacity)
    ]


def _validate_xlsx_rows(frames, matrix):
    for name, frame in frames.items():
        if len(frame) + 1 > _MAX_XLSX_ROWS:
            raise ValueError("%s exceeds Excel's row limit." % name)
    # Two headers, one blank row, the legend title, then each interaction type.
    matrix_rows = len(matrix) + 4 + len(VALID_TYPES)
    if matrix_rows > _MAX_XLSX_ROWS:
        raise ValueError("Residue Matrix exceeds Excel's row limit.")


def _format_standard_sheet(worksheet):
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = _fill("#003B5C")
        cell.alignment = Alignment(horizontal="center")
    for column in range(1, worksheet.max_column + 1):
        values = [worksheet.cell(row, column).value for row in range(1, min(worksheet.max_row, 200) + 1)]
        width = max((len(str(value)) for value in values if value is not None), default=8)
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 42)


def _style_workbook(workbook):
    for name in ("Summary", "Detail", "Parameters", "Input QC"):
        _format_standard_sheet(workbook[name])
    summary = workbook["Summary"]
    summary_header = {cell.value: cell.column for cell in summary[1]}
    for kind in VALID_TYPES:
        column = summary_header.get(kind)
        if column:
            summary.cell(1, column).fill = _fill(color_hex(kind))
    detail = workbook["Detail"]
    detail_header = {cell.value: cell.column for cell in detail[1]}
    type_column = detail_header.get("interaction_type")
    if type_column:
        for row in range(2, detail.max_row + 1):
            kind = detail.cell(row, type_column).value
            if kind in INTERACTION_COLORS:
                detail.cell(row, type_column).fill = _fill(color_hex(kind))
    qc = workbook["Input QC"]
    qc_header = {cell.value: cell.column for cell in qc[1]}
    status_column = qc_header.get("status")
    if status_column:
        status_colors = {"error": "#D55E00", "warning": "#E69F00", "pending": "#F0E442", "success": "#009E73"}
        for row in range(2, qc.max_row + 1):
            color = status_colors.get(qc.cell(row, status_column).value)
            if color:
                qc.cell(row, status_column).fill = _fill(color)


def _neutralize_formulas(workbook):
    """The DockLens schema emits no formulas; force formula-looking input to text."""
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    value = cell.value or ""
                    cell.value = "'" + value
                    cell.data_type = "s"


def export_xlsx(result, path, export_filter=None) -> str:
    """Write the five-sheet DockLens schema v2 workbook atomically."""
    export_filter = export_filter or ExportFilter()
    view = build_export_view(result, export_filter)
    output = Path(os.fspath(path))
    if output.suffix.lower() != ".xlsx":
        output = output.with_name(output.name + ".xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    matrix = residue_matrix_dataframe(view, mode=export_filter.matrix_mode)
    frames = {
        "Summary": summary_dataframe(view),
        "Detail": detail_dataframe(view),
        "Parameters": parameters_dataframe(result, export_filter),
        "Input QC": input_qc_dataframe(result),
    }
    _validate_xlsx_rows(frames, matrix)
    handle = tempfile.NamedTemporaryFile(
        suffix=".xlsx", prefix=".docklens-", dir=output.parent, delete=False
    )
    tmp_path = Path(handle.name)
    handle.close()
    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for name, frame in frames.items():
                _sanitize_dataframe(frame).to_excel(writer, sheet_name=name, index=False)
            workbook = writer.book
            matrices = _matrix_chunks(matrix)
            numbered = len(matrices) > 1
            for index, matrix_chunk in enumerate(matrices, 1):
                title = "Residue Matrix %d" % index if numbered else "Residue Matrix"
                _write_matrix_sheet(
                    workbook, matrix_chunk, title=title, position=index
                )
            _style_workbook(workbook)
            _neutralize_formulas(workbook)
        os.replace(tmp_path, output)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    return str(output)


__all__ = [
    "ExportFilter",
    "build_export_view",
    "detail_dataframe",
    "export_csv",
    "export_xlsx",
    "input_qc_dataframe",
    "parameters_dataframe",
    "residue_matrix_dataframe",
    "summary_dataframe",
]
